from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

wb = Workbook()

# ── COLOUR PALETTE ──────────────────────────────────────────────────────────
DARK     = "1C1C2E"   # dark navy (C&S brand)
ORANGE   = "E8611A"   # C&S orange
GREEN    = "2E7D32"   # 911 Junk green
WHITE    = "FFFFFF"
LGRAY    = "F5F5F5"
MGRAY    = "E0E0E0"
YELLOW   = "FFF9C4"

def hdr_font(color=WHITE, sz=11, bold=True):
    return Font(name="Arial", bold=bold, color=color, size=sz)

def cell_font(sz=10, bold=False, color="000000"):
    return Font(name="Arial", size=sz, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def wrap_align(h="left", v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)

START_DATE = datetime(2026, 3, 23)  # Monday after current date

# ─────────────────────────────────────────────────────────────────────────────
# POST CONTENT
# ─────────────────────────────────────────────────────────────────────────────

# Each post: (day_offset, post_type, ig_fb_caption, tiktok_script, hashtags, visual_direction)
# day_offset: 0=Week1Mon, 1=Week1Tue, etc.

CSD_POSTS = [
  # WEEK 1
  (0, "📸 Project",
   "Before → After. This is what a full interior gut looks like.\n\nThis homeowner needed everything stripped down to the studs before their renovation contractor could start — kitchen, bathrooms, flooring, all of it. We showed up Monday morning and had it done by Tuesday afternoon.\n\n✓ All debris hauled same day\n✓ Site broom-clean\n✓ Permit handled by us\n\nIf you've got a demo project coming up, call for a free on-site estimate.\n📞 (562) 204-6335 | cnsdemo.com\nLicensed CA Contractor #1126325",
   "HOOK (text on screen): 'We gutted this entire house in 2 days 👀'\nSHOW: Before clip of old kitchen/bathrooms → crew arriving → walls coming down → debris loaded → clean empty shell\nON-SCREEN TEXT throughout: city name, service, phone number\nEND CARD: C&S Demolition | (562) 204-6335 | cnsdemo.com",
   "#demolitioncontractor #interiordemolition #orangecounty #losangeles #renovation #beforeandafter #homeremodel #licensed #socal #demolition",
   "Before/after split of interior gut. Show empty shell at end. Tag city."),

  (1, "📚 Educational",
   "Selective demolition vs. full gut — here's the difference 🔨\n\nSELECTIVE DEMO: Only specific elements come out — a wall here, flooring there, one bathroom. The rest of the structure stays. Perfect for targeted remodels.\n\nFULL GUT: Everything goes — walls, ceilings, flooring, fixtures — down to the studs. Best for full renovations or flips.\n\nNot sure which one your project needs? Call us. We'll walk the site and tell you exactly what has to go — no charge.\n\n📞 (562) 204-6335 | cnsdemo.com",
   "HOOK: 'Selective demo vs. full gut — which one do you actually need? 🤔'\nFORMAT: Talking head or text-on-screen bullets\nBULLET 1: Selective = targeted removal, structure stays\nBULLET 2: Full gut = down to studs, everything out\nBULLET 3: 'Not sure? We'll tell you for free'\nEND: Phone number + handle",
   "#demolition #renovation #selectivedemolition #contractor #orangecounty #homeimprovement #remodel #demolitiontips #losangeles #socal",
   "Side-by-side graphic or clips: one wall removed (selective) vs. empty shell (full gut)"),

  (2, "📸 Project",
   "This pool hadn't been used in 10 years. Now it's a backyard.\n\nCracked, unfenced, a safety hazard — and this homeowner in [City] was done with it. We removed the entire structure, backfilled, and compacted the area. Now it's flat usable yard that adds real livable space.\n\nPool removal in Southern California typically runs $3,500–$15,000 depending on size and access. We handle all permits.\n\n📞 Free estimate: (562) 204-6335\ncnsdemo.com | CA License #1126325",
   "HOOK: 'Removing a pool nobody wanted 🚜'\nSHOW: Neglected pool → excavator arriving → concrete breaking → debris hauled → flat clean yard\nON-SCREEN: 'Pool Removal in [City], CA' → cost range → 'Free estimate: (562) 204-6335'\nEND CARD: C&S Demolition",
   "#poolremoval #pooldemolition #orangecounty #demolition #backyard #homeimprovement #losangeles #demolitioncontractor #socal #poolfillin",
   "Before: neglected pool. During: breaking concrete. After: flat graded yard. Show all 3 stages."),

  (3, "🎬 Behind the Scenes",
   "7am. [City]. Here's what the first 2 hours on a demo site actually look like.\n\nBefore any wall comes down — utilities confirmed off, hazmat clearance done, adjacent rooms protected. Prep work is what separates licensed contractors from guys with a sledgehammer.\n\nThen we move. Fast, clean, and safe.\n\n📞 (562) 204-6335 | cnsdemo.com\n✓ CA Licensed #1126325 ✓ Fully Insured",
   "HOOK: '7am on a demo site. Watch what we do first 👀'\nSHOW: Crew truck pulling up → walk-through of site → protection going down → utilities check → first wall comes down\nKEEP: Fast-paced, satisfying cuts\nON-SCREEN: Step labels ('Step 1: Utility check', etc.)\nEND: Business name + number",
   "#demolition #behindthescenes #contractor #orangecounty #crew #construction #demolitionday #licensed #socal #demolitioncontractor",
   "Crew arriving, site walkthrough, first demo action. Raw, authentic, energetic."),

  (4, "📚 Educational",
   "Do you need a permit for demolition in Orange County? Here's the real answer.\n\nThe short version: usually yes.\n\n✓ Removing a non-load-bearing wall in a remodel — sometimes no permit\n✓ Tearing down a detached garage or shed — YES\n✓ Pool demolition — YES\n✓ Full interior gut — depends on scope\n✓ Any structural element — YES\n\nEvery city in OC has slightly different rules. Anaheim, Irvine, Santa Ana, Huntington Beach — we know them all. C&S Demolition pulls all permits on your behalf.\n\n📞 (562) 204-6335 | cnsdemo.com",
   "HOOK: 'Do you need a permit to demo that wall? Let me explain 🏗️'\nFORMAT: Fast bullet-point talking head or text on screen\nCOVER: Shed = yes, Pool = yes, Non-structural wall = sometimes, Full gut = usually yes\nEND: 'We handle all permits. Call for a free estimate.' + number",
   "#demolitionpermit #permits #renovation #orangecounty #contractor #losangeles #homeimprovement #constructionpermit #demolition #socal",
   "Photo of permit paperwork, or quick graphic checklist. Could also be a talking head to camera."),

  (5, "📸 Project",
   "2,000 sq ft of concrete. Gone in one day.\n\nThis driveway and backyard slab in [City] had to go before a new addition could go in. We broke it up, loaded it out, and graded the site — all in one shift.\n\nContractor was on site the next morning.\n\nConcrete removal in SoCal: $1,000–$6,000 depending on thickness and access.\n\n📞 Free estimate: (562) 204-6335 | cnsdemo.com",
   "HOOK: 'Removing 2,000 sq ft of concrete in one day 🔨'\nSHOW: Jackhammer action → excavator breaking → trucks loading → clean graded site\nSATISFYING: Keep cuts tight on the breaking action — this performs well\nEND: Service + city + phone number",
   "#concreteremoval #demolition #contractor #orangecounty #losangeles #concretebreaking #driveway #slab #demolitioncontractor #socal",
   "Jackhammer/excavator close-ups. Timelapse of truck loading if possible. Clean site reveal at end."),

  (6, "🔥 Promo",
   "Free estimate. We come to you. Same week.\n\nIf you've got a demo project — or you're not even sure what you need yet — call us. We walk the site, tell you exactly what it involves, and give you a written quote the same day. No pressure. No obligation.\n\nC&S Demolition is a CA-licensed, fully insured demolition contractor serving Orange County, Los Angeles, Riverside, and San Bernardino Counties.\n\n📞 (562) 204-6335\ncnsdemo.com | License #1126325",
   "HOOK: 'One call. Free estimate. Same week. Here's how it works 📞'\nSHOW: Branded truck → crew walking a job site → written quote → finished project\nKEEP: Clean, professional, trust-building\nEND: 'Call (562) 204-6335' — hold on screen for 3 seconds",
   "#demolition #freeestimate #orangecounty #losangeles #demolitioncontractor #licensed #insured #renovation #contractor #socal",
   "Clean branded shot of truck + crew. Or a collage of finished jobs. Professional feel."),

  # WEEK 2
  (7, "📸 Project",
   "Detached garage. Concrete slab included. Done in 2 days.\n\nThis [City] homeowner wanted the old garage gone to make room for an ADU. We demolished the structure, hauled everything out, and broke up the slab — site was clear and ready for permits.\n\nGarage demolition in SoCal: $2,500–$8,000 depending on size.\n\n📞 Free estimate: (562) 204-6335\ncnsdemo.com | CA License #1126325",
   "HOOK: 'Old garage to flat lot in 48 hours 🏗️'\nSHOW: Old garage standing → crew tearing it down → slab breaking → empty flat lot\nON-SCREEN: 'Making room for an ADU in [City]'\nEND: C&S Demolition + phone",
   "#garagedemolition #demolition #adu #orangecounty #losangeles #contractor #demolitioncontractor #garageremoval #socal #homeimprovement",
   "Before shot of standing garage → action shots of teardown → clean flat lot after."),

  (8, "📚 Educational",
   "How much does interior demolition actually cost in Southern California?\n\nHere's the honest breakdown 👇\n\n🔹 Single bathroom gut: $800–$3,500\n🔹 Kitchen demo: $1,500–$5,000\n🔹 One-room interior: $1,500–$4,000\n🔹 Full house interior gut: $5,000–$15,000+\n\nWhat affects the price:\n✓ Square footage\n✓ Permit requirements\n✓ Hazmat (lead, asbestos)\n✓ Site access for trucks\n✓ Debris disposal distance\n\nThe best way to get an accurate number? A free on-site estimate. We come to you.\n📞 (562) 204-6335 | cnsdemo.com",
   "HOOK: 'How much does interior demo cost in SoCal? Here's the real number 💰'\nFORMAT: Text on screen with quick cuts\nSHOW: Cost tiers scrolling → factors that affect price → CTA\nKEEP: Under 45 seconds\nEND: 'Free estimate — we come to you'",
   "#interiordemolition #demolitioncost #renovation #orangecounty #losangeles #homeimprovement #contractor #remodel #demolition #socal",
   "Cost breakdown graphic or text-on-screen video. Could use before/after of work to illustrate value."),

  (9, "📸 Project",
   "Full bathroom teardown in [City]. Down to the studs and subfloor.\n\nTile, fixtures, vanity, shower walls — all out. Subfloor had water damage so that came out too. Site was clean and ready for the tile contractor by end of day.\n\nBathroom demo in SoCal: $800–$3,500. Same-day turnaround on most jobs.\n\n📞 (562) 204-6335 | cnsdemo.com",
   "HOOK: 'Gutting a bathroom in one day 🚿'\nSHOW: Old bathroom → tile removal → fixtures out → walls stripped → clean shell\nSATISFYING: Tile removal is visually satisfying — get close-up shots\nEND: Service + phone number",
   "#bathroomdemolition #bathroomrenovation #demolition #tileremoval #orangecounty #losangeles #contractor #remodel #demolitioncontractor #socal",
   "Tile smashing close-up, vanity removal, clean empty shell reveal. Very satisfying content."),

  (10, "🎬 Behind the Scenes",
   "Here's what happens to your debris after demo day.\n\nIt doesn't just go to a landfill — at least not all of it.\n\n♻️ Concrete → crushed and recycled for road base\n🪵 Clean wood → sent to wood recyclers or donation\n🔩 Metal → sorted and sent to scrap dealers\n🗑️ Non-recyclable → licensed disposal facility\n\nWe haul it all and handle responsible disposal on every job. That's what a licensed contractor does.\n\n📞 (562) 204-6335 | cnsdemo.com",
   "HOOK: 'What actually happens to your demo debris? 👀'\nSHOW: Debris pile after demo → sorting in truck → concrete crushing facility → clean site\nEDUCATION: Quick text overlays on each material type\nEND: 'C&S Demolition handles all of it'",
   "#demolition #recycling #debrisremoval #contractor #orangecounty #construction #demolitioncontractor #sustainable #socal #losangeles",
   "Show the sorting and loading process. Recycling angle is great for shareability."),

  (11, "📚 Educational",
   "Licensed demolition contractor vs. a handyman with a sledgehammer.\n\nHere's what you actually get with a licensed CA contractor 👇\n\n✓ CSLB license — you can verify it at cslb.ca.gov\n✓ General liability insurance (protects your property)\n✓ Workers' comp (you're not liable if someone gets hurt)\n✓ Permit knowledge — we know what's required in your city\n✓ Hazmat awareness — we won't disturb lead or asbestos unknowingly\n\nThe risk with unlicensed work: if something goes wrong, it's on you. Your homeowner's insurance can deny claims for unpermitted work.\n\nDon't risk it. Call a licensed crew.\n📞 (562) 204-6335 | cnsdemo.com | License #1126325",
   "HOOK: 'Why you shouldn't hire a handyman for demolition ⚠️'\nFORMAT: Talking head or text bullets\nLIST: 5 things licensed contractors have that handymen don't\nKEEP: Informative, not fear-mongering — factual tone\nEND: License number callout",
   "#licensed #contractor #demolition #orangecounty #losangeles #cslb #homeimprovement #renovation #demolitioncontractor #socal",
   "Side-by-side: licensed crew with equipment vs. generic 'beware' graphic. Or just talking head."),

  (12, "📸 Project",
   "Old deck. Rotted boards. Safety hazard. Gone.\n\n[City] homeowner had a deck that was 20+ years old and falling apart. We pulled the whole thing — deck, pergola, and concrete footings — and had the backyard cleared in a day.\n\nDeck demo in SoCal: $800–$3,500 depending on size and material.\n\n📞 Free estimate: (562) 204-6335\ncnsdemo.com | CA License #1126325",
   "HOOK: 'Removing a 20-year-old rotted deck in one day 🔨'\nSHOW: Old sagging deck → boards being pried up → structure coming down → clean empty yard\nEND: Before/after reveal",
   "#deckremoval #deckdemo #demolition #contractor #orangecounty #losangeles #backyard #homeimprovement #demolitioncontractor #socal",
   "Show the rot/damage first (motivates the removal), then the satisfying teardown, then clean yard."),

  (13, "🔥 Promo",
   "Every 5-star review we get comes from one thing: showing up and doing the job right.\n\nLicensed. Insured. Permitted. On time. Site left clean.\n\nIf you've used us before and haven't left a review — it takes 30 seconds and means the world to a small local business.\n\nLink in bio 🙏\n\n📞 (562) 204-6335 | cnsdemo.com\nCA License #1126325",
   "HOOK: 'Every review we get comes from one thing 🙏'\nSHOW: Quick montage of completed jobs — clean sites, happy results\nEND: 'If you've worked with us, leave us a Google review. Link in bio.'\nKEEP: Genuine, humble tone",
   "#reviews #smallbusiness #contractor #demolition #orangecounty #losangeles #google #localcontractor #demolitioncontractor #socal",
   "Montage of clean job sites and finished projects. Warm, authentic feel. Not salesy."),

  # WEEK 3
  (14, "📸 Project",
   "Old shed. Concrete slab. Same day.\n\nThis backyard shed in [City] had been sitting unused for years. We had the structure demolished, slab broken up, and everything hauled out before 4pm.\n\nSite is now ready for a new structure, landscaping, or just open yard space.\n\nShed demo in SoCal: $500–$2,500. Most are same-day jobs.\n\n📞 (562) 204-6335 | cnsdemo.com",
   "HOOK: 'Taking down a backyard shed in under 4 hours ⏱️'\nSHOW: Old shed standing → quick teardown → slab breaking → hauled and gone → open yard\nSATISFYING: Keep the pace fast — this content performs well under 30 seconds",
   "#shedremoval #sheddemolition #demolition #contractor #orangecounty #losangeles #backyard #sameday #demolitioncontractor #socal",
   "Fast, satisfying teardown. Show the structure coming down quickly. Great short-form content."),

  (15, "📚 Educational",
   "5 signs your remodel needs demolition before construction can start 🏗️\n\n1️⃣ You're gutting a kitchen or bathroom (contractor needs a clean shell)\n2️⃣ You're building an addition where a structure currently sits\n3️⃣ The existing layout doesn't match the new floor plan\n4️⃣ There's water damage, mold, or rot behind the walls\n5️⃣ Load-bearing walls need to be repositioned\n\nMost general contractors prefer a clean slate before they start. That's where we come in.\n\nFree demo estimate → contractor comes in → your project starts on time.\n\n📞 (562) 204-6335 | cnsdemo.com",
   "HOOK: '5 signs you need a demo crew before your remodel starts 🔨'\nFORMAT: Quick numbered list — text on screen with short clips of each scenario\nKEEP: Punchy. One point per second.\nEND: 'Free estimate. We come to you.'",
   "#remodel #demolition #renovation #orangecounty #contractor #homeimprovement #losangeles #demolitioncontractor #socal #interiordesign",
   "Quick montage: each scenario illustrated with a clip or photo. Educational carousel could work on IG too."),

  (16, "📸 Project",
   "Commercial interior demo in [City]. Office build-out prep.\n\nFull interior teardown of a commercial suite — ceilings, partitions, flooring — cleared and ready for the new tenant build-out in 3 days.\n\nC&S Demolition handles commercial projects throughout Orange County and LA. Tight timelines, after-hours work available.\n\n📞 (562) 204-6335 | cnsdemo.com\nCA License #1126325",
   "HOOK: 'Turning a commercial space back to shell in 3 days 🏢'\nSHOW: Existing office interior → ceiling tiles coming down → partitions removed → flooring out → clean open shell\nCOMMERCIAL ANGLE: Great for B2B reach",
   "#commercialdemolition #interiordemolition #commercial #contractor #orangecounty #losangeles #officedemolition #demolition #demolitioncontractor #socal",
   "Before: furnished/existing office. After: clean open shell. Show the transformation."),

  (17, "🎬 Behind the Scenes",
   "Here's how the permit process actually works when you hire us.\n\nStep 1: We assess your project and identify which permits are required in your city.\nStep 2: We prepare and submit all permit applications on your behalf.\nStep 3: We coordinate inspections and keep you updated.\nStep 4: Work is signed off and you have a legal record of the project.\n\nYou don't have to figure any of this out. We've done it in every city across OC and LA.\n\n📞 (562) 204-6335 | cnsdemo.com",
   "HOOK: 'How the permit process works when you hire C&S Demo 📋'\nFORMAT: Step-by-step breakdown — text on screen or talking head\n4 STEPS: Assess → Apply → Coordinate → Sign-off\nEND: 'You don't have to figure this out. Call us.'",
   "#permits #demolitionpermit #contractor #orangecounty #losangeles #construction #homeimprovement #licensed #demolition #socal",
   "Show permit paperwork, city building department signage, or a simple step graphic."),

  (18, "📚 Educational",
   "What to expect on demo day 📋\n\nIf you've never hired a demolition contractor before, here's exactly what happens:\n\n🕖 6-8am: Crew arrives, walk the site with you\n🔧 Hour 1-2: Protection laid, utilities confirmed off, hazmat check done\n💥 Hour 2+: Demo begins — structured, not chaotic\n🚛 Throughout: Debris loaded as we go\n🧹 End of day: Site broom-clean, walkthrough with you\n\nNo surprises. We tell you the plan before we start.\n\n📞 (562) 204-6335 | cnsdemo.com",
   "HOOK: 'What actually happens on demo day — here's the full timeline 📋'\nFORMAT: Hour-by-hour breakdown with clips to match each stage\nKEEP: Reassuring tone — great for first-time customers\nEND: 'Call for a free estimate. No surprises.'",
   "#demolition #demoday #contractor #orangecounty #renovation #homeimprovement #demolitioncontractor #losangeles #construction #socal",
   "Chronological clips of a real job day. Crew arriving → work → clean exit. Very reassuring content."),

  (19, "📸 Project",
   "Driveway removal and grading in [City].\n\nThis 40-foot concrete driveway had to go before a new pour could go in. We broke it up, hauled it out, and graded the base — contractor poured the new slab the next day.\n\nConcrete driveway removal in SoCal: $800–$4,000.\n\n📞 Free estimate: (562) 204-6335\ncnsdemo.com | CA License #1126325",
   "HOOK: 'Removing a concrete driveway so a new one can go in 🚛'\nSHOW: Existing cracked driveway → jackhammer/saw cutting → excavator scooping → haul out → clean graded base\nEND: Before/after reveal",
   "#drivewayremoval #concreteremoval #driveway #demolition #contractor #orangecounty #losangeles #concrete #demolitioncontractor #socal",
   "Show the cracked original driveway, the breaking process, and the clean base ready for new pour."),

  (20, "🔥 Promo",
   "CA Licensed #1126325. Here's why that number matters.\n\nWhen you hire C&S Demolition, you can verify our license at cslb.ca.gov. That number means:\n\n✓ We're bonded — your property is protected\n✓ We carry liability insurance — we cover damage if anything goes wrong\n✓ We carry workers' comp — you're not liable for our crew\n✓ We're legally allowed to pull permits in your city\n\nAnyone can show up with a truck. Not everyone can prove they're licensed.\n\n📞 (562) 204-6335 | cnsdemo.com",
   "HOOK: 'CA License #1126325 — here's what it actually means for you 🔒'\nFORMAT: Credentials breakdown — text on screen, punchy\nLIST: 4 protections the license provides\nEND: 'Verify us at cslb.ca.gov — we'll wait'",
   "#licensed #cslb #contractor #demolition #orangecounty #losangeles #bonded #insured #demolitioncontractor #socal",
   "Clean text graphic with license number prominently displayed. Or talking head with confident delivery."),

  # WEEK 4
  (21, "📸 Project",
   "Pool fill-in in [City]. Better value than full removal.\n\nNot every pool needs to be fully demolished. This homeowner wanted the pool gone but had a tight budget — we did a partial demolition and fill-in at $3,500 vs $9,000 for full removal.\n\nThe pool is now safe, compliant, and the yard is usable. Full disclosure given to the buyer on title.\n\n📞 We'll tell you which option makes sense for your situation: (562) 204-6335",
   "HOOK: 'Full pool removal vs. fill-in — which one should you do? 💰'\nFORMAT: Quick comparison. Cost difference, pros/cons\nSHOW: Fill-in process — concrete breaking, fill material going in, compaction, topsoil\nEND: 'Call us — we'll tell you which makes sense'",
   "#poolremoval #poolfillin #pooldemolition #demolition #orangecounty #contractor #losangeles #backyard #demolitioncontractor #socal",
   "Show the fill-in process — it looks different than full removal. Educational and visually interesting."),

  (22, "📚 Educational",
   "Load-bearing wall vs. non-load-bearing wall removal — what's the difference? 🏗️\n\nNON-LOAD-BEARING: The wall only holds itself up. Removing it is relatively straightforward — demo the drywall, studs, and track. Usually no structural work needed.\n\nLOAD-BEARING: The wall carries the weight of the floor or roof above. Removing it requires a temporary support wall, a structural engineer to spec the beam, and a permit. More involved — but very doable.\n\nC&S Demolition handles both. We assess the wall type during your free estimate.\n\n📞 (562) 204-6335 | cnsdemo.com",
   "HOOK: 'Load-bearing vs. non-load-bearing wall removal — here's what matters 🔨'\nFORMAT: Side-by-side explanation. Drawings or clips.\nEDUCATION: This is highly searched content — do a thorough breakdown\nEND: 'We assess both. Free estimate.'",
   "#wallremoval #loadbearingwall #renovation #demolition #orangecounty #contractor #homeimprovement #losangeles #remodel #socal",
   "Diagram or clip showing the difference. Talking head also works well for this educational topic."),

  (23, "📸 Project",
   "Kitchen and bathroom demo — same day.\n\nThis [City] homeowner was doing a full ground-floor remodel. We did the kitchen and main bathroom in a single day — everything out, debris hauled, site broom-clean.\n\nTwo rooms. One crew. One day.\n\nLet your contractor start on day 2.\n📞 (562) 204-6335 | cnsdemo.com",
   "HOOK: 'Kitchen AND bathroom gutted in one day 🔨'\nSHOW: Walk through both spaces before → crew working both simultaneously → both cleared by end of day\nKEEP: Fast-paced montage. Show the efficiency.\nEND: 'Two rooms. One day. Free estimate.'",
   "#kitchendemolition #bathroomdemolition #interiordemolition #demolition #orangecounty #losangeles #renovation #remodel #contractor #socal",
   "Split screen or sequential before/afters of both rooms. Show the efficiency of getting it all done at once."),

  (24, "🎬 Behind the Scenes",
   "Start to finish. This is what a complete interior gut actually looks like.\n\nThis [City] property went from fully finished interior to bare studs in 2 days. Kitchen, 2 bathrooms, all flooring, ceiling drywall in one room — all out.\n\nThe renovation contractor was on site day 3.\n\nThis is the work. Every job, every time.\n📞 (562) 204-6335 | cnsdemo.com",
   "HOOK: 'Full interior gut — start to finish 👀'\nFORMAT: Timelapse or fast-cut montage of full job\nSHOW: Finished interior → crew arriving → everything coming out → clean shell\nMUSIC: High energy background track\nEND: 'C&S Demolition | (562) 204-6335'",
   "#demolition #interiordemolition #beforeandafter #timelapse #orangecounty #losangeles #contractor #renovation #demolitioncontractor #socal",
   "Best content format: timelapse or rapid-cut montage from beginning to end of a full gut job."),

  (25, "📚 Educational",
   "Chimney demolition — what you need to know before you start.\n\nChimneys look simple but they're one of the more technical demo jobs 🧱\n\n✓ Interior chimney vs. exterior — very different process\n✓ Chimney might share a wall with a neighbor (if townhome/condo)\n✓ Fireplace hearth often goes on a separate permit\n✓ Asbestos in older fireplaces is common — always test before demo\n✓ Structural assessment needed if the chimney is load-bearing\n\nWe assess every chimney job before we start. Free on-site estimate.\n📞 (562) 204-6335 | cnsdemo.com",
   "HOOK: 'Thinking about removing your chimney? Watch this first 🧱'\nFORMAT: Educational breakdown — what makes chimney demo tricky\nSHOW: Chimney exterior, interior hearth, brick coming down\nKEEP: Informative, safety-forward tone\nEND: 'Free estimate. We assess everything first.'",
   "#chimneydemolition #chimneyremoval #demolition #contractor #orangecounty #losangeles #fireplace #homeimprovement #demolitioncontractor #socal",
   "Show a chimney exterior and interior hearth. Video of bricks coming down is very satisfying."),

  (26, "📸 Project",
   "Mobile home demolition in [City]. Site cleared and ready.\n\nFull mobile home and manufactured home removal — utilities disconnected, structure demolished, foundation broken up, site graded. Everything out in 3 days.\n\nMobile home demo in SoCal: $3,000–$12,000 depending on size and utility disconnection.\n\n📞 Free estimate: (562) 204-6335\ncnsdemo.com | CA License #1126325",
   "HOOK: 'Removing an entire mobile home in 3 days 🚛'\nSHOW: Mobile home standing → utilities disconnect → demolition begins → debris hauled → clean graded site\nSCALE: Show the size of the job — mobile home demo is impressive content",
   "#mobilehomedemolition #demolition #contractor #orangecounty #losangeles #mobilehome #demolitioncontractor #siteprep #socal #construction",
   "Before: mobile home on site. During: structure coming down. After: clean graded land. Full transformation."),

  (27, "🔥 Promo",
   "We serve 125+ cities across Southern California.\n\nOrange County. Los Angeles County. Riverside County. San Bernardino County.\n\nFrom Anaheim to Lancaster. Irvine to Rancho Cucamonga. Santa Ana to Pomona.\n\nWherever your project is — we're coming to you. Free estimate. Same week.\n\n📞 (562) 204-6335\ncnsdemo.com | CA License #1126325",
   "HOOK: 'C&S Demolition serves 125+ cities in SoCal — here's the map 🗺️'\nSHOW: Montage of job site clips from different cities, each labeled\nOR: Animated map of service area\nEND: Full county list + phone number",
   "#demolition #orangecounty #losangeles #riverside #sanbernardino #contractor #socal #demolitioncontractor #southerncalifornia #licensed",
   "City/county montage works great here. Label each clip with the city name. Shows geographic reach."),
]

# ─── 911 JUNK CA ─────────────────────────────────────────────────────────────
JCA_POSTS = [
  # WEEK 1
  (0, "📸 Project",
   "Before → After. This is what a garage cleanout looks like.\n\nThis [City] garage had years of accumulated junk — furniture, appliances, boxes, scrap. We cleared the entire thing in under 3 hours.\n\n✓ Everything hauled\n✓ Site swept clean\n✓ Donation items separated and dropped off\n\nSame-day service available across Los Angeles and Orange County.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Clearing a garage that hasn't been touched in years 🚛'\nSHOW: Packed garage → crew carrying stuff out → truck loading → empty clean garage\nSATISFYING: This content goes viral. Keep it fast-paced.\nEND: '911 Junk CA | 562-204-6335'",
   "#junkremoval #garagecleanout #orangecounty #losangeles #cleanout #beforeandafter #junk #hauling #sameday #socal",
   "Overstuffed garage → empty clean space. The more dramatic the before, the better the content."),

  (1, "📚 Educational",
   "How much does junk removal actually cost in Orange County?\n\nHere's the honest breakdown 💰\n\n🔹 Single item (couch, mattress, appliance): $75–$150\n🔹 Truckload (1/4 full): $125–$200\n🔹 Truckload (half full): $225–$325\n🔹 Full truck: $400–$600\n🔹 Garage cleanout: $300–$700\n🔹 Estate cleanout: $600–$2,000+\n\nPricing is based on volume and weight — not by the hour. You pay for what we take.\n\nFree quotes always. 📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Junk removal cost in OC — here's the real number 💰'\nFORMAT: Price breakdown — text on screen or talking head\nKEEP: Transparent pricing builds trust. This performs well.\nEND: 'Free quote. No surprises.'",
   "#junkremoval #junkremovalcost #orangecounty #losangeles #pricing #junk #cleanout #hauling #socal #junkca",
   "Price breakdown graphic or text-on-screen. Transparency on pricing is a major trust signal."),

  (2, "📸 Project",
   "Furniture removal in [City]. Old sectional, mattress, and dresser — all gone in one trip.\n\nWe loaded everything, separated what could be donated, and dropped off at a local donation center before the rest went to disposal.\n\nSingle-item or full-room removals welcome. Same-day availability.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Old furniture out of your house in under an hour 🛋️'\nSHOW: Bulky furniture in room → crew carrying out → truck loaded → empty room\nDONATION ANGLE: Show items going to donation — people love this\nEND: '911 Junk CA | Same-day service'",
   "#furnitureremoval #junkremoval #orangecounty #losangeles #mattressremoval #donation #cleanout #sameday #junkca #socal",
   "Show the furniture first (relatable — everyone has old furniture), then the quick removal process."),

  (3, "🎬 Behind the Scenes",
   "Here's what we do with your stuff before it hits the landfill.\n\nNot everything goes to the dump. Every load we pick up gets sorted 👇\n\n♻️ Metal → scrap yard\n🪑 Furniture in good condition → donation center\n📦 Reusable goods → Goodwill or Habitat for Humanity ReStore\n🗑️ Actual junk → licensed disposal facility\n\nWe try to divert as much as possible. It's better for you, better for the planet.\n\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'We don't just dump everything — here's what actually happens 👀'\nSHOW: Truck being loaded → sorting process → donation center drop-off → recycling yard\nWHY: This content builds massive goodwill. People care about this.\nEND: '911 Junk CA — we donate & recycle first'",
   "#junkremoval #donation #recycling #orangecounty #losangeles #goodwill #sustainableliving #junkca #socal #hauling",
   "Show the sorting and donation drop-off. Very shareable content — eco-friendly angle resonates strongly."),

  (4, "📚 Educational",
   "Junk removal vs. renting a dumpster — which one should you choose?\n\nHere's how to decide 👇\n\n🚛 JUNK REMOVAL (us):\n✓ We load it for you\n✓ No permit needed\n✓ No minimum rental period\n✓ Done in hours, not days\n✓ Better for mixed loads\n\n🗑️ DUMPSTER RENTAL:\n✓ Better for large construction debris\n✓ You load it yourself\n✓ Requires driveway/street permit in most OC cities\n✓ Rental period = pressure to fill it fast\n\nFor most home cleanouts, junk removal is faster and easier. For demolition debris or week-long projects, a dumpster might make more sense.\n\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Junk removal vs. dumpster rental — which one wins? 🤔'\nFORMAT: Side-by-side comparison. Quick, decisive tone.\nKEEP: Under 60 seconds. People want the answer fast.\nEND: 'For most cleanouts — call us. Same day, no loading.'",
   "#junkremoval #dumpsterrental #orangecounty #losangeles #cleanout #junk #homeimprovement #socal #junkca #hauling",
   "Simple comparison graphic or talking head. This is a top-searched question — great for traffic."),

  (5, "📸 Project",
   "Estate cleanout in [City]. Full house cleared in 2 days.\n\nAfter a family member passes, the last thing anyone wants to deal with is a house full of belongings. We helped this family clear the entire property — furniture, appliances, personal items, garage — with care and efficiency.\n\nEstate cleanouts: $600–$2,000+ depending on size.\n\nWe handle it with respect. That's the job.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Estate cleanout — we handle the hard part 🏠'\nTONE: Respectful and warm. This is an emotional topic.\nSHOW: Full furnished house → crew working methodically → empty clean property\nMESSAGE: 'We give families one less thing to worry about'\nEND: Soft CTA — '911 Junk CA. Call when you're ready.'",
   "#estatecleanout #junkremoval #cleanout #orangecounty #losangeles #estatesale #hoarding #hauling #junkca #socal",
   "Keep it respectful. Show organized, careful removal — not chaotic. The tone matters here."),

  (6, "🔥 Promo",
   "Same-day junk removal in Orange County and Los Angeles.\n\nCall before noon → we're there today.\n\n✓ No hidden fees\n✓ Free upfront quote\n✓ We do all the loading\n✓ Donate & recycle what we can\n\nServing 93+ cities across Southern California 🌴\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Call before noon. Junk gone by tonight. 📞'\nFORMAT: Clean, punchy promo. Simple message delivered fast.\nSHOW: Phone call → truck arriving → junk loaded → clean space\nEND: '562-204-6335' — hold 3 seconds\nENERGY: Upbeat, fast-moving",
   "#junkremoval #sameday #orangecounty #losangeles #junk #cleanout #hauling #socallife #junkca #socal",
   "Fast-moving montage: phone, truck, loading, clean result. Keep it under 20 seconds."),

  # WEEK 2
  (7, "📸 Project",
   "Hot tub removal in [City]. These weigh 800 lbs. We handle it.\n\nHot tub had been sitting disconnected on a concrete pad for 4 years. We broke it down on-site, hauled out the pieces, and removed the concrete pad at the same time.\n\nHot tub removal in SoCal: $300–$600 average.\n\nFree quotes. Same-week availability.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Removing an 800-lb hot tub that nobody wanted anymore 💪'\nSHOW: Hot tub on pad → team breaking it down → pieces hauled out → clean pad (then optional: pad removal)\nSATISFYING: People love specialty removals. This performs very well.\nEND: Weight callout + price range + phone number",
   "#hottubremoval #junkremoval #orangecounty #losangeles #hottubremoving #heavyitem #cleanout #junkca #sameday #socal",
   "Show the size of the hot tub first, then the breakdown process. The 800-lb stat is a great hook."),

  (8, "📚 Educational",
   "Items we take that most junk removal companies won't 👀\n\n🎹 Piano (upright or baby grand)\n🎱 Pool table (full disassembly + haul)\n🔒 Fireproof safe (even full ones)\n💪 Gym equipment (treadmills, weight sets, machines)\n🛁 Cast iron tubs (heavy — we have the crew)\n🌿 Yard waste and green waste\n📺 TVs and electronics (responsible e-waste disposal)\n\nOne call. We take it all.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Items we take that other junk companies refuse 👀'\nFORMAT: List reveal — one item at a time, with clips or photos\nKEEP: Each item gets 2-3 seconds. Quick and satisfying.\nEND: 'One call. We take it all. 562-204-6335'",
   "#junkremoval #pianoremoval #pooltableremoval #orangecounty #losangeles #heavyitem #junk #cleanout #junkca #socal",
   "Show each specialty item being removed. The more unusual, the better. People love niche content."),

  (9, "📸 Project",
   "Hoarder cleanout in [City]. We do this with 100% respect and zero judgment.\n\nEvery home is different. Every family situation is different. We show up, we work efficiently, and we treat every person and every job with dignity.\n\nFull home cleanout completed over 2 days.\n\nIf you or someone you know needs help with this — we're here.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Hoarder cleanout — done with respect, no judgment 🙏'\nTONE: Compassionate and professional. Do NOT make this sensational.\nSHOW: Before → methodical organized removal → progress shots → clean home\nMESSAGE: Dignity and efficiency. This builds massive trust.\nEND: Soft CTA with phone number",
   "#hoardercleanout #junkremoval #cleanout #orangecounty #losangeles #hoarding #estatesale #respect #junkca #socal",
   "Show organized, methodical work — NOT exploitation of the situation. Respectful tone throughout."),

  (10, "🎬 Behind the Scenes",
   "Our truck holds more than you think. Here's the breakdown 📦\n\n🚛 Full truck = approximately 10–13 cubic yards\nThat's roughly:\n\n✓ 5–7 full sofas\n✓ 3 full bedroom sets\n✓ OR one completely packed 2-car garage\n✓ OR a mix of appliances, boxes, furniture, and miscellaneous junk\n\nWe quote by volume — you only pay for the space you use. Minimum load pricing starts under $125.\n\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'How much fits in our junk removal truck? 📦'\nSHOW: Empty truck bed → items going in → visual of how much fits → full loaded truck\nVISUAL: Great for showing value — people are always curious about this\nEND: Volume chart + pricing callout",
   "#junkremoval #truckload #orangecounty #losangeles #pricing #hauling #junk #cleanout #junkca #socal",
   "Show truck being loaded with variety of items. Side-by-side of empty vs full is a great comparison shot."),

  (11, "📚 Educational",
   "How to prepare for junk removal day — make it faster and save money 💰\n\n✓ Separate what you want to keep vs. what goes — before we arrive\n✓ Point out donation items (we'll drop them off)\n✓ Clear a path to large items if possible\n✓ Know your HOA rules for truck parking (most are fine)\n✓ Have a list of items ready — we can quote faster\n\nThe more organized your side is, the faster we work — and the less time it takes.\n\n📞 562-204-6335 | 911junkca.com",
   "HOOK: '3 things to do before your junk removal crew shows up ✅'\nFORMAT: Quick checklist — practical and actionable\nKEEP: Short. 30-45 seconds max.\nEND: 'Do these 3 things. We handle the rest.'",
   "#junkremoval #tips #cleanout #orangecounty #losangeles #junk #hauling #homeimprovement #junkca #socal",
   "Simple checklist format works great — text on screen or quick talking head."),

  (12, "📸 Project",
   "Yard waste removal in [City]. After the storm, before the HOA notices.\n\nThis backyard took a beating — broken branches, pile of old pavers, dead plants, yard debris. We had it all cleared and hauled out same day.\n\nYard waste and green waste removal available throughout OC and LA.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Backyard cleanup after the storm 🌿'\nSHOW: Messy debris-filled yard → crew clearing → loaded truck → clean yard\nTIMING: Great for post-storm seasonal content\nEND: 'Same-day green waste removal. 562-204-6335'",
   "#yardwaste #greenwaste #junkremoval #orangecounty #losangeles #yarddebris #cleanup #sameday #junkca #socal",
   "Before/after of messy yard vs. clean yard is very satisfying. Works well on all platforms."),

  (13, "🔥 Promo",
   "We serve 93+ cities across Los Angeles and Orange County.\n\nFrom Anaheim to Long Beach. Irvine to Torrance. Santa Ana to Culver City.\n\nSame-day service in most areas. Call before noon, done today.\n\n✓ No hidden fees\n✓ Free upfront quote\n✓ We do all the loading\n\n📞 562-204-6335 | 911junkca.com",
   "HOOK: '93+ cities. Same-day service. One call. 📞'\nSHOW: City name montage or map visual\nKEEP: Short and punchy — this is a reach play\nEND: Phone number held for 3 seconds\nFEEL: Confident, established",
   "#junkremoval #orangecounty #losangeles #sameday #socal #junk #cleanout #hauling #junkca #southerncalifornia",
   "City/area montage with city names labeled. Or a simple map animation of service area."),

  # WEEK 3
  (14, "📸 Project",
   "Appliance graveyard in [City]. Fridge, washer, dryer, dishwasher — all in one trip.\n\nThese were left behind after a tenant move-out. We cleared all 4 appliances, plus the miscellaneous junk left in the unit, in under 2 hours.\n\nAppliance removal starts at $75/item. Multi-item discounts available.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: '4 appliances. 2 hours. One trip. Let's go 🚛'\nSHOW: Appliances waiting in unit → crew strapping and loading → truck loaded → clean empty unit\nRELATABLE: Landlords and property managers love this content\nEND: Per-item pricing callout",
   "#applianceremoval #junkremoval #orangecounty #losangeles #fridge #washerdryerremoval #moveout #landlord #junkca #socal",
   "Show the appliances lined up first (relatable for landlords), then the quick efficient removal."),

  (15, "📚 Educational",
   "E-waste — what actually happens to your old electronics? ♻️\n\nPlease don't throw them in the trash. Here's why — and what we do instead.\n\n📱 Phones, tablets, laptops → data-wiped, sorted, refurbished or recycled\n📺 TVs and monitors → contain lead — must go to certified e-waste recyclers\n🖥️ Computers → components recovered (copper, gold, rare metals)\n🔋 Batteries → hazardous if landfilled — require special disposal\n\nCalifornia law bans most electronics from regular trash. We handle responsible e-waste disposal on every job.\n\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'What actually happens to your old electronics? ♻️'\nFORMAT: Educational breakdown — this topic gets a lot of engagement\nSHOW: Old TV/computer pile → e-waste facility → sorting process\nCA LAW CALLOUT: Good local angle\nEND: 'We handle it right. Call us.'",
   "#ewaste #electronics #junkremoval #recycling #orangecounty #losangeles #tv #computers #ewasterecycling #socal",
   "Show a pile of old electronics, then the responsible recycling process. Educational angle = shareable."),

  (16, "📸 Project",
   "Office cleanout in [City]. Business relocated — left everything behind.\n\nFull office clear-out: desks, chairs, filing cabinets, equipment, breakroom appliances — everything loaded in one day.\n\nWe work around business hours and can do after-hours cleanouts to minimize disruption.\n\nCommercial cleanouts: pricing by volume. Free quote.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Clearing out an entire office in one day 🏢'\nSHOW: Full office with furniture → crew stripping it out → loaded trucks → empty office ready for next tenant\nB2B ANGLE: Target property managers and business owners\nEND: 'Commercial cleanouts. After-hours available.'",
   "#officecleanout #junkremoval #commercial #orangecounty #losangeles #office #business #hauling #junkca #socal",
   "Before: furnished office. After: empty, clean. Show the scale of a commercial job."),

  (17, "🎬 Behind the Scenes",
   "Garage cleanout timelapse. Watch 20 years of stuff disappear in 60 seconds.\n\nThis is probably our most satisfying video format. A completely packed garage, cleared start to finish.\n\nThis [City] garage had: old furniture, sports equipment, boxes of mystery stuff, broken appliances, and a riding lawnmower.\n\n2 crew members. 2.5 hours. Gone.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Garage cleanout timelapse — watch this whole thing 👀'\nFORMAT: Timelapse of full garage cleanout from start to finish\nKEEP: This is your highest-performing content format. Nail the execution.\nMUSIC: Upbeat track\nEND: Before/after freeze frame + phone number",
   "#junkremoval #garagecleanout #timelapse #orangecounty #losangeles #garage #cleanout #satisfying #junkca #socal",
   "Timelapse from first item removed to empty clean garage. This format consistently goes viral."),

  (18, "📚 Educational",
   "Move-out cleanout checklist — what to do with everything you're not taking 📋\n\n✓ Furniture too bulky to move — junk removal (us)\n✓ Clothes and linens → Goodwill, Salvation Army, or shelters\n✓ Small appliances in good condition → Facebook Marketplace or donation\n✓ Kitchen items → Habitat for Humanity ReStore\n✓ Mattresses → most cities have bulk item pickup (1x/month)\n✓ Electronics → e-waste drop-off or junk removal with e-waste service\n✓ Hazardous materials (paint, chemicals) → city HHW drop-off only\n\nFor everything else — one call to us.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Moving out? Here's what to do with everything you can't take 📦'\nFORMAT: Category-by-category breakdown\nKEEP: Practical and actionable. People are stressed when moving — this helps.\nEND: 'For everything else — 562-204-6335'",
   "#moveout #moving #junkremoval #cleanout #orangecounty #losangeles #movingday #movingtips #junkca #socal",
   "Moving boxes/chaos imagery. Checklist format works well as Instagram carousel too."),

  (19, "📸 Project",
   "Shed and backyard cleanout in [City]. Old shed + 10 years of backyard accumulation.\n\nShed hauled away, broken patio furniture gone, old landscaping equipment out — yard cleared in one visit.\n\nWe take sheds (we break them down on-site), old yard equipment, and anything else cluttering your outdoor space.\n\nFree quote.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Backyard hoarder shed situation — cleared in one visit 🌿'\nSHOW: Cluttered backyard/shed → crew working → truck loaded → clean open yard\nVERSATILE: Appeal to homeowners prepping for summer\nEND: 'One visit. Free quote.'",
   "#shedremoval #backyardcleanout #junkremoval #orangecounty #losangeles #shed #backyard #cleanup #junkca #socal",
   "Show the before backyard cluttered state — the messier, the better for engagement. Then the clean result."),

  (20, "🔥 Promo",
   "Leaving a Google review takes 30 seconds. And it means everything to us.\n\nIf 911 Junk CA showed up, did the job right, and left your space clean — we'd be incredibly grateful for a quick review.\n\nLink in bio. 🙏\n\nIf you haven't used us yet — call for a free quote and let's change that.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: '30 seconds. That's all it takes. 🙏'\nFORMAT: Genuine, heartfelt ask. No gimmicks.\nSHOW: Clean job results montage\nMESSAGE: Authentic small-business appeal\nEND: 'Link in bio for Google review + 562-204-6335 for a quote'",
   "#junkremoval #googlereview #smallbusiness #orangecounty #losangeles #review #cleanout #junkca #socal #hauling",
   "Warm montage of satisfied results. Genuine appeal — not a sales pitch."),

  # WEEK 4
  (21, "📸 Project",
   "Piano removal in [City]. Yes, we move pianos. ♟️\n\nUpright piano from a second-floor bedroom — we get asked about this constantly. The answer is yes, we do it, and no, you don't want to try it yourself.\n\nPiano removal: $200–$500 depending on size and access.\n\nOne call. We figure out the logistics.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Removing a piano from a second-floor bedroom 🎹 (this is harder than it looks)'\nSHOW: Piano in room → crew strapping it → careful staircase navigation → loaded in truck\nCHALLENGE ANGLE: People love watching a tricky removal go smoothly\nEND: 'Piano removal in SoCal — 562-204-6335'",
   "#pianoremoval #junkremoval #orangecounty #losangeles #piano #heavyitem #specialty #cleanout #junkca #socal",
   "The staircase moment is the money shot. Show the careful process — people are impressed by the skill."),

  (22, "📚 Educational",
   "What we cannot take — and why 🚫\n\nWe take almost everything. But here's what legally can't go in our truck:\n\n⛽ Gas, oil, and flammable liquids\n🧪 Pesticides, herbicides, pool chemicals\n🔋 Car batteries (accepted at AutoZone/O'Reilly free)\n💊 Medications (take to Walgreens or CVS drop-off)\n🎨 Paint (liquid — take to city HHW event or PaintCare drop-off)\n☢️ Asbestos or lead materials\n\nFor everything else — call us. If we can't take it, we'll tell you exactly where to bring it.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'What junk removal companies legally CANNOT take 🚫'\nFORMAT: List reveal with icons or clips\nWHY: People always want to know this. High-value informational content.\nEND: 'Everything else — we take it. Call us.' + drop-off alternatives",
   "#junkremoval #hazardouswaste #ewaste #orangecounty #losangeles #tips #junk #cleanout #junkca #socal",
   "Simple list graphic or talking head. Link to city resources in caption for bonus engagement."),

  (23, "📸 Project",
   "Pool table removal in [City]. This thing weighed 900 lbs.\n\nDisassembled on-site, pieces wrapped and loaded, slate separated for recycling. Gone in 2 hours.\n\nPool tables are one of the most searched specialty removals in Southern California. We've done hundreds.\n\nPool table removal: $250–$500 average.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Removing a 900-lb pool table — here's how it's done 🎱'\nSHOW: Pool table in place → disassembly → slate removal → loaded and hauled\nWEIGHT CALLOUT: The 900-lb detail is a great hook\nEND: Price + phone",
   "#pooltableremoval #junkremoval #orangecounty #losangeles #pooltable #heavyitem #specialty #junkca #socal #cleanout",
   "Show the disassembly process — people are curious how it's done. The slate weighs 400+ lbs alone."),

  (24, "🎬 Behind the Scenes",
   "End of day recap. Here's what the crew got done today.\n\n9am: Garage cleanout in Irvine — 3 hours\n1pm: Hot tub removal in Anaheim — 1.5 hours\n3:30pm: Move-out cleanout in Fullerton — 2 hours\n\nThree jobs. One day. Every space left clean.\n\nThat's 911 Junk CA.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'What our crew did today — 3 jobs, one day 🚛'\nFORMAT: Job-by-job recap. Quick clips or stills from each location.\nSHOWS: Efficiency and volume capacity. Trust-building.\nCITY LABELS: Name each city — good for local discovery\nEND: 'Book your spot. 562-204-6335'",
   "#junkremoval #orangecounty #losangeles #irvine #anaheim #fullerton #behindthescenes #crew #junkca #socal",
   "Multiple job clips from one day — show the team busy and productive. Creates FOMO/demand signal."),

  (25, "📚 Educational",
   "Same-day junk removal — is it actually possible? ✅\n\nYes. Here's how it works:\n\n📞 You call or text us before noon\n📍 We confirm your city is on today's route\n🚛 We give you a 2-hour arrival window\n✅ We show up, load everything, and you're done\n\nSame-day availability depends on route coverage and job volume. We cover most of Orange County and LA daily.\n\nCall first thing in the morning for the best availability.\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Same-day junk removal — is it actually real? Let me explain ✅'\nFORMAT: Step-by-step process. Reassuring and clear.\nKEEP: Under 45 seconds. This is a high-intent search topic.\nEND: 'Call before noon. Done today.' + number",
   "#junkremoval #sameday #orangecounty #losangeles #sameDayService #junk #cleanout #hauling #junkca #socal",
   "Simple process explainer — show the call-to-completion timeline. Very practical and conversion-focused."),

  (26, "📸 Project",
   "Full home cleanout after an estate sale in [City].\n\nEverything that didn't sell stayed behind. We cleared 3 bedrooms, living room, kitchen, and garage — all in one day.\n\nPost-estate sale cleanouts are a specialty. We work fast, work with families respectfully, and leave the property ready for the realtor.\n\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'Post-estate sale cleanout — what gets left behind 🏡'\nTONE: Respectful and professional\nSHOW: Remaining items post-sale → efficient removal → clean empty house\nAUDIENCE: Families dealing with estate, realtors, property managers\nEND: 'Ready for the realtor. Same day.'",
   "#estatecleanout #estatesale #junkremoval #orangecounty #losangeles #realtor #cleanout #junkca #socal #hauling",
   "Show the post-sale reality — items left behind after buyers have gone. Then the clean empty result."),

  (27, "🔥 Promo",
   "No hidden fees. No surprise charges. Free upfront quote.\n\nHere's our promise: we give you a price before we start. If it changes, we tell you why before we move anything.\n\nYou pay for what we take. Nothing more.\n\n✓ Same-day service available\n✓ 93+ cities in OC and LA\n✓ We donate & recycle first\n✓ Free quotes always\n\n📞 562-204-6335 | 911junkca.com",
   "HOOK: 'No hidden fees. We give you a price before we touch anything. 💰'\nFORMAT: Transparency/trust play. Simple, clean delivery.\nSHOW: Quote being given → crew working → clean result\nEND: '562-204-6335 — free quote. No commitment.'\nFEEL: Honest, no-pressure",
   "#junkremoval #nohiddenfees #transparent #orangecounty #losangeles #freeestimate #junk #cleanout #junkca #socal",
   "Clean branded content. The no-hidden-fees message is a strong differentiator — lead with it."),
]

# ─────────────────────────────────────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────

HEADERS = ["Day", "Date", "Post Type", "Caption (Instagram & Facebook)", 
           "TikTok Script", "Hashtags", "Visual Direction / Notes"]

COL_WIDTHS = [6, 12, 14, 60, 45, 38, 35]

def build_sheet(wb, title, posts, brand_color, tab_color):
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A2"

    # Header row
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font()
        cell.fill = fill(brand_color)
        cell.alignment = wrap_align("center", "center")
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22

    for day_offset, ptype, caption, tiktok, hashtags, visual in posts:
        row_num = day_offset + 2
        date = START_DATE + timedelta(days=day_offset)
        week = (day_offset // 7) + 1
        day_of_week = date.strftime("%a")
        date_str = date.strftime("%b %-d")
        bg = LGRAY if (day_offset // 7) % 2 == 0 else WHITE

        row_data = [
            f"W{week} {day_of_week}",
            date_str,
            ptype,
            caption,
            tiktok,
            hashtags,
            visual
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = wrap_align()
            cell.border = thin_border()
            cell.fill = fill(bg)
            cell.font = cell_font()
            if col == 3:  # Post type column — bold
                cell.font = cell_font(bold=True)
                if "📸" in val:
                    cell.fill = fill("E3F2FD")
                elif "📚" in val:
                    cell.fill = fill("F3E5F5")
                elif "🎬" in val:
                    cell.fill = fill("E8F5E9")
                elif "🔥" in val:
                    cell.fill = fill("FFF3E0")

        ws.row_dimensions[row_num].height = 90

# Build sheets
wb.remove(wb.active)  # remove default blank sheet
build_sheet(wb, "C&S Demolition", CSD_POSTS, DARK, "E8611A")
build_sheet(wb, "911 Junk CA", JCA_POSTS, GREEN, "2E7D32")

# ── INSTRUCTIONS SHEET ───────────────────────────────────────────────────────
ws_info = wb.create_sheet("How to Use", 0)
ws_info.sheet_properties.tabColor = "1565C0"
ws_info.column_dimensions["A"].width = 90

instructions = [
    ("HOW TO USE THIS SOCIAL MEDIA CALENDAR", True, "1565C0", WHITE, 14),
    ("", False, WHITE, "000000", 11),
    ("PLATFORMS: Post every caption to Instagram AND Facebook (same text works for both). Use the TikTok Script column for video content.", False, "E3F2FD", "000000", 11),
    ("SCHEDULE: Use Buffer, Later, or Meta Business Suite (free) to schedule IG and FB posts in advance. TikToks must be posted manually or via TikTok's native scheduler.", False, "E3F2FD", "000000", 11),
    ("", False, WHITE, "000000", 11),
    ("POST TYPES & WHAT THEY MEAN", True, "37474F", WHITE, 12),
    ("📸 Project Posts (3x per week) — Before/after of real jobs. Tag the city. Always include service and phone number.", False, "E3F2FD", "000000", 11),
    ("📚 Educational Posts (2x per week) — Tips, cost guides, how-to info. These build trust and get saved/shared.", False, "F3E5F5", "000000", 11),
    ("🎬 Behind the Scenes (1x per week) — Raw, authentic content from the job site. Crew, process, day-in-the-life.", False, "E8F5E9", "000000", 11),
    ("🔥 Promo Posts (1x per week) — CTA, credentials, service area, review requests. Keep it concise.", False, "FFF3E0", "000000", 11),
    ("", False, WHITE, "000000", 11),
    ("TIKTOK TIPS", True, "37474F", WHITE, 12),
    ("✓ Hook in the first 1-2 seconds is everything — use the exact hook text from the TikTok Script column", False, "E3F2FD", "000000", 11),
    ("✓ Keep videos under 60 seconds for max reach (30-45 seconds is the sweet spot)", False, "E3F2FD", "000000", 11),
    ("✓ Always add on-screen text — most people watch without sound", False, "E3F2FD", "000000", 11),
    ("✓ Post between 6-9am or 7-10pm for best algorithm performance", False, "E3F2FD", "000000", 11),
    ("✓ Reply to every comment in the first hour after posting — boosts reach", False, "E3F2FD", "000000", 11),
    ("", False, WHITE, "000000", 11),
    ("INSTAGRAM/FACEBOOK TIPS", True, "37474F", WHITE, 12),
    ("✓ Always tag the city in both the caption and as a location tag when you post", False, "E8F5E9", "000000", 11),
    ("✓ Post Reels (video) whenever you have footage — they get 3-5x more reach than static photos", False, "E8F5E9", "000000", 11),
    ("✓ If you only have photos, post a carousel (multiple photos) — higher engagement than single image", False, "E8F5E9", "000000", 11),
    ("✓ Best times: Tuesday-Friday, 9-11am or 6-8pm", False, "E8F5E9", "000000", 11),
    ("✓ Respond to every DM and comment — the algorithm rewards engagement", False, "E8F5E9", "000000", 11),
    ("", False, WHITE, "000000", 11),
    ("FREE SCHEDULING TOOLS", True, "37474F", WHITE, 12),
    ("Meta Business Suite (free) — Schedule IG + FB posts and Reels. Access at business.facebook.com", False, "FFF9C4", "000000", 11),
    ("Buffer (free plan) — Schedule across IG, FB, TikTok, LinkedIn. buffer.com", False, "FFF9C4", "000000", 11),
    ("Later (free plan) — Great visual calendar for Instagram. later.com", False, "FFF9C4", "000000", 11),
    ("TikTok Creator Studio — Schedule TikToks natively for free. studio.tiktok.com", False, "FFF9C4", "000000", 11),
    ("", False, WHITE, "000000", 11),
    ("WHERE TO FILL IN [CITY] — Replace every [City] placeholder with the actual city where the job was done.", True, "B71C1C", WHITE, 11),
]

for r, (text, bold, bg_hex, fg_hex, sz) in enumerate(instructions, 1):
    cell = ws_info.cell(row=r, column=1, value=text)
    cell.font = Font(name="Arial", bold=bold, size=sz, color=fg_hex)
    cell.fill = fill(bg_hex)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_info.row_dimensions[r].height = 28 if bold else 22

output_path = "/sessions/gifted-youthful-mendel/mnt/cns-demolition-pseo/Social Media Content Calendar - C&S Demo + 911 Junk CA.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
